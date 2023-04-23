import pandas as pd
import openpyxl
import json
import random


def import_tables(filepath):
    """
    This function takes a json listing the desired encounter tables to use from
    an excel workbook and returns the list.
    :param filepath: filepath
    :return: list of str
    """
    try:
        with open(filepath, "r") as fp:
            content = fp.read()
    except IOError:
        print(f"{filepath} not found or is not readable.")
        raise Exception(f"Configuration Problem: {filepath} was not readable.")

    raw_data = json.loads(content)
    return raw_data["table list"]


def import_workbook(tabs: list, filepath):
    """
    This function takes a list of tab names and a filepath to an excel workbook
    and returns a dictionary of pandas dataframes, using the tab names as keys.
    The first two columns are created by transforming the D100 string columns into
    separate columns of integers labeled Roll and Max.
    :param tabs: list of str
    :param filepath: filepath
    :return: dict of pd.DataFrame
    """
    wb = pd.read_excel(filepath, sheet_name=tabs, index_col=None, na_values=False)
    for tab in tabs:
        print(f"tab: {tab}")
        rolls = []
        maxes = []
        for i in range(len(wb[tab])):
            item = wb[tab]["D100"][i].split("–")
            # print(f"item: {item}")
            if len(item) > 1:
                rolls.append(int(item[0]))
                maxes.append(int(item[1]))
            else:
                try:
                    temp = int(item[0])
                except ValueError:
                    item = wb[tab]["D100"][i].split("-")
                rolls.append(int(item[0]))
                maxes.append(int(item[0]))
        roll_s = pd.Series(rolls, index=None)
        max_s = pd.Series(maxes, index=None)
        wb[tab]["Roll"] = roll_s
        wb[tab]["Max"] = max_s
    return wb


def roll_result(table: pd.DataFrame) -> str:
    """
    This function accepts a random encounter table, determines the range for the
    random roll, and returns the string result corresponding to the roll. This table
    must have the following columns: Roll (int), Max (int), TYPE (str), and
    ENCOUNTER (str). There can be no overlap between the ranges specified by Roll
    and Max in each row and Roll <= Max. There can be no blank spaces in these columns.
    :param table: pd.DataFrame
    :return: str: result
    """
    # Find the min and max for the roll range. These encounter tables can have
    # numbers far above 100.
    min_roll = table["Roll"].min()
    max_roll = table["Max"].max()

    roll = random.randint(min_roll, max_roll)
    print(f"roll: {roll}")
    # Filter for smaller than or equal to roll in table.Roll.
    filtered = table.loc[table.Roll <= roll]
    print(f"1st filter: {filtered}")
    # Now, filter again for larger than or equal to roll in table.Max.
    # This should reduce it to a single row if there are no overlaps.
    filtered = filtered.loc[filtered.Max >= roll]
    print(f"2nd filter: {filtered}")
    result = filtered["ENCOUNTER"].squeeze()
    type_result = filtered["TYPE"].squeeze()
    return f"{result} ({type_result})"


def validate_workbook(tables, filepath):
    """
    This function looks at each of the tables (DataFrames) in the workbook to
    make sure that all of them adhere to the required standards. It will return
    a list of bad tables if any are found or None if the workbook is ok.
    The requirements are as follows:
        1. All worksheets (ws) must contain columns titled D100, ENCOUNTER, and TYPE.
        2. All fields must be strings.
        3. D100 must contain characters that read as positive integer, but are still
            strings.
        4. In D100, each field must be either a single "integer" or a pair of
            "integers" separated by a dash - or – character (due to print formatting).
        5. In D100, the first integer must be less or equal to the second integer.
        6. In D100, the integers must be listed in ascending order. There is no
            maximum, however.
        7. In D100, no range of integers may overlap any other, including singles.
        8. In D100, there can be no gaps of even in single number either.
    :param tables: list of str
    :param filepath: filepath
    :return: list of str or None
    """
    wb = pd.read_excel(filepath, sheet_name=tables, index_col=None, na_values=False)
    print(f"Loaded basic workbook")
    bad_tables = []
    for table in tables:
        print(f"Testing table: {table}")
        ws = wb[table]
        print(f"Worksheet: {ws}")
        # Testing the first property.
        try:
            roll_series = ws["D100"]
            enc_series = ws["ENCOUNTER"]
            type_series = ws["TYPE"]
        except KeyError:
            bad_tables.append(f"{table} Property 1 Failed")
            continue
        # print(f"{table} passed Property 1 Test.")

        # Testing the second property.
        ds = ws.applymap(lambda x: isinstance(x, str)).all()
        if ds.all() is False:
            bad_tables.append(f"{table} Property 2 Failed")
            continue
        # print(f"{table} passed Property 2 Test.")

        # The import_workbook function will test properties 3 and 4.
        try:
            new_wb = import_workbook([table], filepath)
        except ValueError:
            bad_tables.append(f"{table} Property 3 and/or 4 Failed")
            continue
        else:
            new_ws = new_wb[table]
        # print(f"{table} passed Property 3 and 4 Test.")

        # Testing property 5.
        ds = new_ws["Roll"] <= new_ws["Max"]
        if ds.all() is False:
            bad_tables.append(f"{table} Property 5 Failed")
            continue
        # print(f"{table} passed Property 5 Test.")

        # Testing property 6.
        r1 = r2 = new_ws["Roll"].tolist()
        m1 = m2 = new_ws["Max"].tolist()
        r1.sort()
        m1.sort()
        if r1 != r2 or m1 != m2:
            bad_tables.append(f"{table} Property 6 Failed")
            continue
        # print(f"{table} passed Property 6 Test.")

        # Finally, we check properties 7 and 8. We know that both series ascend
        # and Roll will never be greater than Max. Every roll result between the
        # minimum roll and maximum roll must produce exactly one row. Overlaps
        # will produce more than one result, skips will produce an empty dataframe.
        min_roll = new_ws["Roll"].min()
        max_roll = new_ws["Max"].max()
        bad = False
        bad_nums = []
        for i in range(min_roll, max_roll + 1):
            filtered = new_ws.loc[new_ws.Roll <= i]
            filtered = filtered.loc[filtered.Max >= i]
            rows = filtered.shape[0]
            print(f"Rows for Index {i}: {rows}")
            if rows != 1:
                bad = True
                bad_nums.append(str(i))
        if bad is True:
            bad_tables.append(f"{table} Prop 7/8 Failed at indices: {bad_nums}")
            continue
        print(f"{table} passed all tests.")

    if not bad_tables:
        return None
    else:
        return bad_tables


if __name__ == "__main__":
    tables = import_tables("./samples/tables.json")
    print(tables)
    bad_tables = validate_workbook(tables, "./samples/encounters.xlsx")
    if bad_tables is None:
        print("Workbook passed the tests.")
        workbook = import_workbook(tables, "./samples/encounters.xlsx")
        print(workbook)
        idx = random.randint(0, len(tables) - 1)
        print(f"idx: {idx}")
        table = workbook[tables[idx]]
        print(table)
        result = roll_result(table)
        print(f"result: {result}")
    else:
        print("The following tables are improperly formatted.")
        for table in bad_tables:
            print(table)
        print("The program cannot continue testing itself.")
    next_table_set = import_tables("./data/tables.json")
    next_test = validate_workbook(next_table_set, "./data/Random Encounters.xlsx")
    if next_test is not None:
        print("Workbook failed at least one test.")
        for item in next_test:
            print(item)
